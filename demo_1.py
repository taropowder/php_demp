import sys
from PyQt5.QtWidgets import (QWidget,QLabel,QPushButton,QLineEdit,QInputDialog,QApplication)
from PyQt5.QtWidgets import QGridLayout
import xlrd
import re
import openpyxl
import os

class GUI(QWidget):
    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):


        self.lbl = QLabel('请输入当前周数,\n点击按钮\n就会在当前文件夹下面生成文件',self)
        self.Edit = QLineEdit()
        self.btn = QPushButton('生成文件',self)
        self.lb2 = QLabel('未生成',self)


        grid = QGridLayout()
        grid.setSpacing(10)

        grid.addWidget(self.lbl,1,0)
        grid.addWidget(self.Edit,1,1)
        grid.addWidget(self.btn,2,0)
        grid.addWidget(self.lb2,3,0)
        self.setLayout(grid)
        # print(self.Edit.text())
        self.btn.clicked.connect(self.getinput)

        self.setGeometry(300, 300, 350, 180)
        self.setWindowTitle('无课表生成器')
        # self.setWindowIcon(QIcon('web.png'))
        # self.center()
        self.show()

    def getinput(self):
        week = self.Edit.text()
        excelwork(week)
        self.lb2.setText("已经生成当前目录下第 %s 周无课表" % week)







def changenumber(lessionnumber):
    LessionRegex= re.compile(r'(\d*),')
    LessionList = LessionRegex.findall(lessionnumber)
    NewNumber= []
    if LessionList :
        LessionRegex = re.compile(r'(\d*)周')
        LessionList2 = LessionRegex.findall(lessionnumber)
        LessionList = LessionList + LessionList2
        for j in LessionList:
            NewNumber.append(int(j))
        return NewNumber
    else:
        LessionRegex = re.compile(r'(\d*?)-(\d*)')
        LessionList = LessionRegex.findall(lessionnumber)
        for j in range(int(LessionList[0][0]), int(LessionList[0][1]) + 1):
            NewNumber.append(j)
        return NewNumber




def ReadExcel (excelname):
    workbook = xlrd.open_workbook(excelname)
    sheet = workbook.sheet_by_index(0)
    persion = []
    LessionTime1 = []
    LessionTime2 = []
    LessionTime3 = []
    LessionTime4 = []
    LessionTime5 = []
    name = sheet.cell(26,2).value
    for i in range(2,7):
        LessionTime1.append(sheet.cell(5,i).value)
        LessionTime2.append(sheet.cell(9, i).value)
        LessionTime3.append(sheet.cell(14, i).value)
        LessionTime4.append(sheet.cell(18, i).value)
        LessionTime5.append(sheet.cell(23, i).value)
    for j in range(0,5):
        if LessionTime1[j]!='':
            LessionTime1[j]=changenumber(LessionTime1[j])
        if LessionTime2[j]!='':
            LessionTime2[j]=changenumber(LessionTime2[j])
        if LessionTime3[j]!='':
            LessionTime3[j]=changenumber(LessionTime3[j])
        if LessionTime4[j]!='':
            LessionTime4[j]=changenumber(LessionTime4[j])
        if LessionTime5[j] != '':
            LessionTime5[j]=changenumber(LessionTime5[j])
    persion.append(LessionTime1)
    persion.append(LessionTime2)
    persion.append(LessionTime3)
    persion.append(LessionTime4)
    persion.append(LessionTime5)

    return persion,name

def writeexcel(persion,name,if_frist):
    if if_frist:
        for row in range(2, 7):
            for columu in range(2, 7):
                if persion[row - 2][columu - 2] == "":
                    sheet.cell(row=row, column=columu).value =name + "&"
                elif week  in persion[row - 2][columu - 2]:
                    continue
                elif week not in persion[row - 2][columu - 2]:
                    sheet.cell(row=row, column=columu).value =name + "&"
    else:
        for row in range(2, 7):
            for columu in range(2, 7):
                if persion[row - 2][columu - 2] == "":
                    sheet.cell(row=row, column=columu).value = str(sheet.cell(row=row, column=columu).value)+name + "&"
                elif week in persion[row - 2][columu - 2]:
                    continue
                else:
                    sheet.cell(row=row, column=columu).value = str(sheet.cell(row=row, column=columu).value)+name + "&"
    return False


def CheckPath(filepath):
    ListFile= []
    for i in os.listdir(filepath):
        if i.endswith('.xls'):
            ListFile.append(i)
    return ListFile

def excelwork(weekend):
    # week = 8
    Filepath = '.'
    global week
    week=int(weekend)
    # week = int(sys.argv[1])
    # Filepath = sys.argv[2]
    wb = openpyxl.Workbook()
    global sheet
    sheet = wb.get_sheet_by_name('Sheet')
    if_frist = True
    sheet['A1'] = '节次'
    sheet['B1'] = '星期一'
    sheet['C1'] = '星期二'
    sheet['D1'] = '星期三'
    sheet['E1'] = '星期四'
    sheet['F1'] = '星期五'
    sheet['A2'] = '第一大节'
    sheet['A3'] = '第二大节'
    sheet['A4'] = '第三大节'
    sheet['A5'] = '第四大节'
    sheet['A6'] = '第五大节'
    FileList = CheckPath(Filepath)
    listlen = len(FileList)
    sheet.row_dimensions[2].height = listlen+18
    sheet.row_dimensions[3].height = listlen+18
    sheet.row_dimensions[4].height = listlen+18
    sheet.row_dimensions[5].height = listlen+18
    sheet.row_dimensions[6].height = listlen+18
    sheet.column_dimensions['B'].width = listlen*8.5
    sheet.column_dimensions['C'].width = listlen*8.5
    sheet.column_dimensions['D'].width = listlen*8.5
    sheet.column_dimensions['E'].width = listlen*8.5
    sheet.column_dimensions['F'].width = listlen*8.5
    # print("欢迎进入无课表制作脚本")
    for filename in FileList:
        persion,name=ReadExcel(filename)
        if_frist=writeexcel(persion,name,if_frist)
        # print("正在导入 %s 课表" % name)
    wb.save('第 %s 周无课表.xlsx'% week)



if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = GUI()
    sys.exit(app.exec_())
